"""social_account/service.py —— 自媒体账号汇总的业务逻辑

- 账号 CRUD、排序、图片上传
- Excel 导入（智能识别列名 + 重复策略）
- Excel 导出（全量 / 选中，纯文本，不含图片列）
- 图片资源导出/导入（自包含 zip：Excel + social_images/ 文件夹，可还原图片）
- 图片关联：zip 包内按 平台+账号ID 稳定命名（social_images/<平台>_<账号ID>_qr.<ext>），
  导入时按该键匹配，不依赖 db id，故清空重建/改 Excel 后图片仍对得上原账号。
- UI 上传图片仍按 db id 命名（{id}_qr.{ext}），覆盖式，移除即删文件。
"""

import io
import os
import re
import shutil
import tempfile
import uuid
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from common import db

IMG_DIR = db.DATA_DIR / "social_images"
ALLOWED_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}

# 文件名不安全字符（含中文/空格/路径符等），用于把账号标识清洗成稳定的图片文件名
_UNSAFE = re.compile(r'[<>:"/\\|?*\x00-\x1f\s]')


def _img_key(platform, base):
    """按 平台+账号ID（无账号ID 时用账号名称）生成稳定的图片关联键，不依赖 db id。"""
    raw = f"{platform}_{base}"
    return _UNSAFE.sub("_", raw or "")[:100]

# 字段顺序即 Excel 列顺序（序号单列在外层）
FIELDS = [
    ("platform", "平台"),
    ("account_name", "账号名称"),
    ("account_id", "账号ID"),
    ("user_id", "UserId"),
    ("homepage_url", "主页链接"),
    ("bio", "简介"),
    ("gender", "性别"),
    ("birthday", "生日"),
    ("location", "所在地"),
    ("likes_count", "获赞"),
    ("mutual_count", "互关"),
    ("following_count", "关注"),
    ("followers_count", "粉丝"),
    ("qr_image", "二维码"),
    ("cover_image", "封面"),
]
INT_FIELDS = {"likes_count", "mutual_count", "following_count", "followers_count"}

# Excel 列名 -> 字段：支持中文 + 英文同义词
COLUMN_ALIASES = {
    "platform": ["平台", "platform", "平台名称"],
    "account_name": ["账号名称", "账号", "名称", "account_name", "name"],
    "account_id": ["账号id", "账号ID", "account_id", "accountid"],
    "user_id": ["userid", "user_id", "UserId", "用户id"],
    "homepage_url": ["主页链接", "主页", "homepage", "homepage_url", "链接"],
    "bio": ["简介", "bio", "description", "签名"],
    "gender": ["性别", "gender", "sex"],
    "birthday": ["生日", "birthday", "出生日期"],
    "location": ["所在地", "地区", "location", "region", "城市"],
    "likes_count": ["获赞", "点赞", "likes", "likes_count", "获赞数"],
    "mutual_count": ["互关", "mutual", "mutual_count", "互关数"],
    "following_count": ["关注", "following", "following_count", "关注数"],
    "followers_count": ["粉丝", "followers", "followers_count", "粉丝数"],
    "qr_image": ["二维码", "qr", "qr_image", "二维码图片", "二维码图"],
    "cover_image": ["封面", "cover", "cover_image", "封面图片", "封面图"],
}
_HEADER_TO_FIELD = {}
for _f, _al in COLUMN_ALIASES.items():
    _HEADER_TO_FIELD[_f.lower()] = _f
    for _a in _al:
        _HEADER_TO_FIELD[_a.strip().lower()] = _f

_INSERT_COLS = (
    "platform, account_name, account_id, user_id, homepage_url, bio, gender, "
    "birthday, location, likes_count, mutual_count, following_count, "
    "followers_count, qr_image, cover_image, sort_order"
)


def _conn():
    return db.get_conn()


def _ensure_img_dir():
    IMG_DIR.mkdir(parents=True, exist_ok=True)


def _to_int(v):
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def _clean(data):
    d = {}
    for f, _ in FIELDS:
        val = data.get(f)
        if f in INT_FIELDS:
            d[f] = _to_int(val)
        else:
            d[f] = (val or "").strip()
    return d


def list_accounts():
    conn = _conn()
    try:
        rows = conn.execute(
            "SELECT * FROM social_account ORDER BY sort_order ASC, id ASC"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def get_account(pid):
    conn = _conn()
    try:
        r = conn.execute("SELECT * FROM social_account WHERE id=?", (pid,)).fetchone()
        return dict(r) if r else None
    finally:
        conn.close()


def create_account(data):
    d = _clean(data)
    conn = _conn()
    try:
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM social_account").fetchone()[0]
        cur = conn.execute(
            f"INSERT INTO social_account({_INSERT_COLS}) VALUES("
            "?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                d["platform"], d["account_name"], d["account_id"], d["user_id"],
                d["homepage_url"], d["bio"], d["gender"], d["birthday"], d["location"],
                d["likes_count"], d["mutual_count"], d["following_count"],
                d["followers_count"], d["qr_image"], d["cover_image"], max_order + 10,
            ),
        )
        conn.commit()
        new_id = cur.lastrowid
        # 把临时上传的图片（uuid 命名）重命名为固定命名 {new_id}_qr/_cover
        fixed = {}
        for slot, col in (("qr", "qr_image"), ("cover", "cover_image")):
            if d.get(col):
                fixed[col] = finalize_image(new_id, slot, d[col])
        if fixed:
            conn.execute(
                "UPDATE social_account SET qr_image=?, cover_image=? WHERE id=?",
                (fixed.get("qr_image", d["qr_image"]),
                 fixed.get("cover_image", d["cover_image"]), new_id),
            )
            conn.commit()
        return get_account(new_id)
    finally:
        conn.close()


def update_account(pid, data):
    d = _clean(data)
    conn = _conn()
    try:
        old = conn.execute(
            "SELECT qr_image, cover_image FROM social_account WHERE id=?", (pid,)
        ).fetchone()
        # 图片：移除则删文件；否则重命名为固定命名（覆盖式）
        for slot, col in (("qr", "qr_image"), ("cover", "cover_image")):
            old_rel = old[col] if old else None
            if not d.get(col):
                if old_rel:
                    _remove_file(old_rel)
            else:
                d[col] = finalize_image(pid, slot, d[col])
        conn.execute(
            "UPDATE social_account SET platform=?, account_name=?, account_id=?, user_id=?, "
            "homepage_url=?, bio=?, gender=?, birthday=?, location=?, likes_count=?, "
            "mutual_count=?, following_count=?, followers_count=?, qr_image=?, cover_image=?, "
            "updated_at=(datetime('now','localtime')) WHERE id=?",
            (
                d["platform"], d["account_name"], d["account_id"], d["user_id"],
                d["homepage_url"], d["bio"], d["gender"], d["birthday"], d["location"],
                d["likes_count"], d["mutual_count"], d["following_count"],
                d["followers_count"], d["qr_image"], d["cover_image"], pid,
            ),
        )
        conn.commit()
        return get_account(pid)
    finally:
        conn.close()


def delete_account(pid):
    conn = _conn()
    try:
        row = conn.execute(
            "SELECT qr_image, cover_image FROM social_account WHERE id=?", (pid,)
        ).fetchone()
        conn.execute("DELETE FROM social_account WHERE id=?", (pid,))
        conn.commit()
        if row:
            for rel in (row["qr_image"], row["cover_image"]):
                if rel:
                    p = db.DATA_DIR / rel
                    if p.exists():
                        try:
                            p.unlink()
                        except Exception:
                            pass
        return {"deleted": pid}
    finally:
        conn.close()


def reorder(ordered_ids):
    conn = _conn()
    try:
        for i, pid in enumerate(ordered_ids):
            conn.execute(
                "UPDATE social_account SET sort_order=? WHERE id=?", ((i + 1) * 10, pid))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


def save_image(file_bytes, filename):
    _ensure_img_dir()
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXT:
        ext = ".png"
    rel = f"social_images/{uuid.uuid4().hex}{ext}"
    (db.DATA_DIR / rel).write_bytes(file_bytes)
    return rel


def get_image_file(rel_path):
    target = (db.DATA_DIR / rel_path).resolve()
    base = IMG_DIR.resolve()
    if not str(target).startswith(str(base)):
        raise ValueError("路径越界")
    if not target.exists():
        raise ValueError("文件不存在")
    return target


def _remove_file(rel):
    """删除某个相对路径对应的图片文件（移除账号/清空图片时调用）。"""
    if not rel:
        return
    try:
        p = db.DATA_DIR / rel
        if p.exists():
            p.unlink()
    except Exception:
        pass


def finalize_image(account_id, slot, rel):
    """把临时上传的图片（uuid 命名）重命名为固定命名 {account_id}_{slot}.{ext}，覆盖式。
    - 若 rel 已是固定命名或对应文件不存在，则原样返回（幂等）。
    - 重命名同时删除该 slot 已有旧固定文件，避免孤儿文件堆积。
    返回最终相对路径。"""
    if not rel:
        return rel
    src = db.DATA_DIR / rel
    if not src.exists():
        return rel
    ext = os.path.splitext(rel)[1].lower()
    if ext not in ALLOWED_EXT:
        ext = ".png"
    new_rel = f"social_images/{account_id}_{slot}{ext}"
    if rel == new_rel:
        return rel
    new_path = db.DATA_DIR / new_rel
    if new_path.exists():
        try:
            new_path.unlink()
        except Exception:
            pass
    try:
        src.rename(new_path)
    except Exception:
        return rel
    return new_rel


def _dup_filter(rec):
    """重复判定：优先 平台+账号ID，账号ID 为空则 平台+账号名称。"""
    pid = rec.get("account_id", "")
    if pid:
        return ("platform=? AND account_id=?", (rec["platform"], pid))
    return ("platform=? AND account_name=?", (rec["platform"], rec.get("account_name", "")))


def _import_rows(grid, mode="skip", image_resolver=None):
    """通用导入行处理。
    image_resolver(field, rec) -> 最终相对路径或 ''（默认忽略图片，仅纯文本导入用）。
    rec 为当前行已解析字段 dict，可用于按 平台+账号ID 关联图片，而不依赖 db id。
    """
    if image_resolver is None:
        image_resolver = lambda field, rec: ""
    if not grid:
        return {"imported": 0, "skipped": 0, "updated": 0, "errors": ["文件为空"]}
    header = grid[0]
    col_map = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = str(h).strip().lower()
        if key in _HEADER_TO_FIELD:
            col_map[idx] = _HEADER_TO_FIELD[key]
    if not col_map:
        raise ValueError("未识别到任何已知列（平台/账号名称/账号ID 等），请检查表头")
    stats = {"imported": 0, "skipped": 0, "updated": 0, "errors": []}
    # 导入仅支持「新增(skip)」与「修改(overwrite)」两种模式，不再提供清空重建
    if mode not in ("skip", "overwrite"):
        mode = "skip"
    conn = _conn()
    try:
        max_order = conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM social_account").fetchone()[0]
        for row_i, r in enumerate(grid[1:], start=2):  # 第 1 行为表头，数据从第 2 行起
            rec = {}
            for idx, field in col_map.items():
                val = r[idx] if idx < len(r) else None
                if field in ("qr_image", "cover_image"):
                    # 图片列：交给 resolver 处理（按 平台+账号ID 关联，纯文本导入返回空 -> 忽略）
                    rec[field] = image_resolver(field, rec)
                    continue
                rec[field] = "" if val is None else str(val).strip()
            if not any(rec.values()):
                continue  # 整行空白跳过
            # 必填校验：平台、账号名称、账号ID 缺一项则该行报错并跳过
            missing = [f for f in ("platform", "account_name", "account_id") if not rec.get(f)]
            if missing:
                _label = {"platform": "平台", "account_name": "账号名称", "account_id": "账号ID"}
                stats["errors"].append(
                    f"第 {row_i} 行缺少必填项：{'、'.join(_label[m] for m in missing)}")
                continue
            df = _dup_filter(rec)
            existing = conn.execute(
                f"SELECT id FROM social_account WHERE {df[0]}", df[1]).fetchone()
            if existing and mode == "skip":
                stats["skipped"] += 1
                continue
            if existing and mode == "overwrite":
                pid = existing["id"]
                d = _clean(rec)
                # 图片列：resolver 返回空（纯文本导入）时保留数据库原图；否则用解析结果
                old = conn.execute(
                    "SELECT qr_image, cover_image FROM social_account WHERE id=?", (pid,)
                ).fetchone()
                if old:
                    if not d["qr_image"]:
                        d["qr_image"] = old["qr_image"] or ""
                    if not d["cover_image"]:
                        d["cover_image"] = old["cover_image"] or ""
                conn.execute(
                    "UPDATE social_account SET platform=?, account_name=?, account_id=?, "
                    "user_id=?, homepage_url=?, bio=?, gender=?, birthday=?, location=?, "
                    "likes_count=?, mutual_count=?, following_count=?, followers_count=?, "
                    "qr_image=?, cover_image=?, updated_at=(datetime('now','localtime')) WHERE id=?",
                    (d["platform"], d["account_name"], d["account_id"], d["user_id"],
                     d["homepage_url"], d["bio"], d["gender"], d["birthday"], d["location"],
                     d["likes_count"], d["mutual_count"], d["following_count"],
                     d["followers_count"], d["qr_image"], d["cover_image"], pid),
                )
                stats["updated"] += 1
                continue
            # mode == "new" 或不存在重复 -> 插入
            d = _clean(rec)
            max_order += 10
            conn.execute(
                f"INSERT INTO social_account({_INSERT_COLS}) VALUES("
                "?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (d["platform"], d["account_name"], d["account_id"], d["user_id"],
                 d["homepage_url"], d["bio"], d["gender"], d["birthday"], d["location"],
                 d["likes_count"], d["mutual_count"], d["following_count"],
                 d["followers_count"], d["qr_image"], d["cover_image"], max_order),
            )
            stats["imported"] += 1
        conn.commit()
    finally:
        conn.close()
    return stats


def import_excel(file_bytes, mode="skip"):
    """纯文本导入：图片列被忽略（resolver 默认返回空）。"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    grid = list(wb.active.iter_rows(values_only=True))
    return _import_rows(grid, mode)


def import_bundle(zip_bytes, mode="skip"):
    """导入自包含压缩包：内含 Excel（图片列为相对路径）+ social_images/ 图片文件夹，可还原图片。"""
    tmp = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            z.extractall(tmp)
        xlsx_path = None
        for root, _dirs, files in os.walk(tmp):
            for f in files:
                if f.lower().endswith(".xlsx"):
                    xlsx_path = os.path.join(root, f)
                    break
            if xlsx_path:
                break
        if not xlsx_path:
            raise ValueError("压缩包内未找到 Excel 文件")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        grid = list(wb.active.iter_rows(values_only=True))

        img_root = os.path.join(tmp, "social_images")

        def resolver(field, rec):
            # 按 平台+账号ID 稳定关联，不依赖 db id（清空重建后 id 变化也能对上）
            slot = "qr" if field == "qr_image" else "cover"
            base = (rec.get("account_id") or rec.get("account_name") or "").strip()
            key = _img_key(rec.get("platform", ""), base)
            if not key:
                return ""
            src = None
            for ext in ALLOWED_EXT:
                cand = os.path.join(img_root, f"{key}_{slot}{ext}")
                if os.path.exists(cand):
                    src = cand
                    break
            if not src:
                return ""
            _ensure_img_dir()
            ext = os.path.splitext(src)[1].lower()
            dst = db.DATA_DIR / "social_images" / f"{key}_{slot}{ext}"
            if dst.exists():
                try:
                    dst.unlink()
                except Exception:
                    pass
            shutil.copy2(src, dst)
            return f"social_images/{key}_{slot}{ext}"

        return _import_rows(grid, mode, resolver)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def export_excel(ids=None, with_images=False):
    conn = _conn()
    try:
        if ids:
            ids = [int(x) for x in ids]
            ph = ",".join("?" * len(ids))
            rows = conn.execute(
                f"SELECT * FROM social_account WHERE id IN ({ph}) "
                "ORDER BY sort_order ASC, id ASC", ids).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM social_account ORDER BY sort_order ASC, id ASC").fetchall()
        rows = [dict(r) for r in rows]
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "自媒体账号"
    headers = ["序号"] + [h for _, h in FIELDS]
    ws.append(headers)
    ws.column_dimensions["A"].width = 6
    for i, (f, _h) in enumerate(FIELDS):
        col = get_column_letter(i + 2)
        ws.column_dimensions[col].width = 18 if f not in ("qr_image", "cover_image") else 14
    ws.freeze_panes = "A2"

    for i, row in enumerate(rows, start=1):
        vals = [i]
        for f, _h in FIELDS:
            v = row.get(f, "")
            if f in INT_FIELDS:
                v = v if v not in (None, "") else 0
            elif f in ("qr_image", "cover_image"):
                # 普通导出留空（避免导入破坏图片）；打包导出带相对路径
                v = row.get(f, "") if with_images else ""
            vals.append(v)
        ws.append(vals)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_bundle(ids=None):
    """导出自包含压缩包：Excel（含图片相对路径）+ social_images/ 图片文件夹。"""
    conn = _conn()
    try:
        if ids:
            ids = [int(x) for x in ids]
            ph = ",".join("?" * len(ids))
            rows = conn.execute(
                f"SELECT * FROM social_account WHERE id IN ({ph}) "
                "ORDER BY sort_order ASC, id ASC", ids).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM social_account ORDER BY sort_order ASC, id ASC").fetchall()
        rows = [dict(r) for r in rows]
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "自媒体账号"
    headers = ["序号"] + [h for _, h in FIELDS]
    ws.append(headers)
    ws.column_dimensions["A"].width = 6
    for i, (f, _h) in enumerate(FIELDS):
        col = get_column_letter(i + 2)
        ws.column_dimensions[col].width = 18 if f not in ("qr_image", "cover_image") else 14
    ws.freeze_panes = "A2"
    # 预计算图片在 zip 内的稳定命名（按 平台+账号ID），不再依赖 db id
    img_map = {}  # (row_idx, col) -> (zip_name, src_path)
    for idx, row in enumerate(rows):
        for col in ("qr_image", "cover_image"):
            rel = row.get(col)
            if not rel:
                continue
            src = db.DATA_DIR / rel
            if not src.exists():
                continue
            base = (row.get("account_id") or row.get("account_name") or "").strip()
            key = _img_key(row.get("platform", ""), base)
            slot = "qr" if col == "qr_image" else "cover"
            ext = os.path.splitext(rel)[1].lower() or ".png"
            zname = f"social_images/{key}_{slot}{ext}"
            img_map[(idx, col)] = (zname, str(src))

    for i, row in enumerate(rows, start=1):
        vals = [i]
        for f, _h in FIELDS:
            v = row.get(f, "")
            if f in INT_FIELDS:
                v = v if v not in (None, "") else 0
            elif f in ("qr_image", "cover_image"):
                # 写出稳定命名路径，便于重新导入时按 平台+账号ID 关联
                v = img_map.get((i - 1, f), ("", None))[0] or ""
            vals.append(v)
        ws.append(vals)
    xls_buf = io.BytesIO()
    wb.save(xls_buf)
    xls_buf.seek(0)

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("账号矩阵.xlsx", xls_buf.getvalue())
        for _zname, _src in img_map.values():
            z.write(_src, _zname)
    out.seek(0)
    return out.getvalue()


def export_template():
    """生成空导入模板（仅表头 + 示例说明）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "自媒体账号"
    headers = ["序号"] + [h for _, h in FIELDS]
    ws.append(headers)
    ws.column_dimensions["A"].width = 6
    for i, (f, _h) in enumerate(FIELDS):
        col = get_column_letter(i + 2)
        ws.column_dimensions[col].width = 18 if f not in ("qr_image", "cover_image") else 14
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
