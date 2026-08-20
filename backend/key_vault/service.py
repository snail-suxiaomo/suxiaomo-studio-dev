"""key_vault/service.py —— AI 密钥库读写（与各厂商 API 密钥/平台信息集中管理）

独立于 model_config：本模块只管「平台/账号/密钥/额度/过期」，模型配置通过 key_vault_id 引用它。
敏感字段（api_key / secret_key）按本地工具处理：接口返回明文，由前端掩码展示。
"""
import io
import os

from common import db
from openpyxl import Workbook, load_workbook

# 全部可写字段（不含 id / 时间戳）
FIELDS = [
    'name', 'provider', 'category', 'base_url', 'api_key', 'secret_key',
    'account', 'dev_url',
]

# 导出 Excel 的列（中文表头 -> 字段）
EXPORT_HEADERS = [
    ('名称', 'name'), ('平台', 'provider'), ('归类', 'category'),
    ('接口地址', 'base_url'), ('API Key', 'api_key'), ('Secret Key', 'secret_key'),
    ('登录账号', 'account'), ('开发者平台', 'dev_url'),
]

# 导入列名同义词 -> 字段
FIELD_SYNONYMS = {
    'name': ['名称', '条目名称', '备注名', '名字', 'name'],
    'provider': ['平台', '厂商', '平台名称', '厂家', '服务商', 'provider'],
    'category': ['归类', '类型', '分类', '类别', 'category'],
    'base_url': ['接口地址', '对接链接', '模型对接链接', 'api地址', 'base_url', 'endpoint', 'api url'],
    'api_key': ['api key', '密钥', 'key', 'apikey', 'token', 'ak', 'api_key'],
    'secret_key': ['secret key', 'secret_key', 'sk', '密钥2', '秘钥'],
    'account': ['账号', '登录账号', '账户', '邮箱', '手机号', 'account'],
    'dev_url': ['开发者平台', '平台网站', '官网', '开发者平台网站', 'console', '控制台', 'dev_url'],
}


def _row_to_dict(r) -> dict:
    d = dict(r)
    return d


def list_vaults(category='全部', keyword=None):
    sql = "SELECT * FROM key_vault WHERE 1=1"
    params = []
    if category and category != '全部':
        sql += " AND category = ?"
        params.append(category)
    if keyword:
        sql += " AND (name LIKE ? OR provider LIKE ? OR account LIKE ?)"
        kw = f"%{keyword}%"
        params.extend([kw, kw, kw])
    sql += " ORDER BY sort_order DESC, id ASC"
    conn = db.get_conn()
    try:
        rows = conn.execute(sql, params).fetchall()
        return [_row_to_dict(r) for r in rows]
    finally:
        conn.close()


def get_vault(vid):
    conn = db.get_conn()
    try:
        row = conn.execute("SELECT * FROM key_vault WHERE id = ?", (vid,)).fetchone()
        return _row_to_dict(row) if row else None
    finally:
        conn.close()


def create_vault(d: dict):
    conn = db.get_conn()
    try:
        max_order = conn.execute("SELECT MAX(sort_order) AS m FROM key_vault").fetchone()["m"] or 0
        cols, vals = [], []
        for f in FIELDS:
            cols.append(f)
            v = d.get(f)
            vals.append(v if v is not None else '')
        cols.append('sort_order')
        vals.append(int(d.get('sort_order', max_order + 10)))
        ph = ','.join('?' * len(cols))
        cur = conn.execute(
            f"INSERT INTO key_vault ({','.join(cols)}) VALUES ({ph})", vals)
        conn.commit()
        return get_vault(cur.lastrowid)
    finally:
        conn.close()


def update_vault(vid, d: dict):
    conn = db.get_conn()
    try:
        exist = conn.execute("SELECT id FROM key_vault WHERE id = ?", (vid,)).fetchone()
        if not exist:
            return None
        sets, vals = [], []
        for f in FIELDS:
            if f in d:
                v = d.get(f)
                sets.append(f"{f} = ?")
                vals.append(v if v is not None else '')
        if sets:
            sets.append("updated_at = datetime('now')")
            conn.execute(
                f"UPDATE key_vault SET {','.join(sets)} WHERE id = ?", vals + [vid])
            conn.commit()
        return get_vault(vid)
    finally:
        conn.close()


def delete_vault(vid):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM key_vault WHERE id = ?", (vid,))
        conn.commit()
        return {'deleted': 1}
    finally:
        conn.close()


def batch_delete(ids: list):
    ids = [int(x) for x in ids]
    if not ids:
        return {'deleted': 0}
    conn = db.get_conn()
    try:
        placeholders = ','.join('?' * len(ids))
        n = conn.execute(
            f"DELETE FROM key_vault WHERE id IN ({placeholders})", ids
        ).rowcount
        conn.commit()
        return {'deleted': n}
    finally:
        conn.close()


def reorder(ids: list):
    conn = db.get_conn()
    try:
        for idx, vid in enumerate(ids):
            conn.execute(
                "UPDATE key_vault SET sort_order = ?, updated_at = datetime('now') WHERE id = ?",
                ((len(ids) - idx) * 10, vid))
        conn.commit()
    finally:
        conn.close()


# ---------------- Excel 导出 / 导入 ----------------

def export_excel(ids=None):
    """导出选中（ids）或全部为 xlsx 字节。注意：含 api_key / secret_key 明文。"""
    conn = db.get_conn()
    try:
        if ids:
            ids = [int(x) for x in ids]
            ph = ','.join('?' * len(ids))
            rows = conn.execute(
                f"SELECT * FROM key_vault WHERE id IN ({ph}) ORDER BY sort_order DESC, id ASC", ids
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM key_vault ORDER BY sort_order DESC, id ASC").fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'AI密钥库'
    ws.append([h for h, _ in EXPORT_HEADERS])
    for r in rows:
        d = dict(r)
        ws.append([d.get(f, '') if d.get(f) is not None else '' for _, f in EXPORT_HEADERS])
    # 列宽自适应（简单按表头长度）
    for col_idx, (h, _) in enumerate(EXPORT_HEADERS, start=1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max(10, len(h) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _match_columns(header_cells):
    """把表头单元格映射到字段名，返回 {列索引: 字段}。"""
    col_map = {}
    for idx, cell in enumerate(header_cells):
        if cell is None:
            continue
        raw = str(cell).strip().lower()
        if not raw:
            continue
        for field, syns in FIELD_SYNONYMS.items():
            if raw in [s.lower() for s in syns]:
                col_map[idx] = field
                break
    return col_map


def import_excel(file_bytes, strategy='skip'):
    """从 xlsx 导入密钥。

    唯一键为组合 (name, provider, account)：
      - 三者均不同则新建
      - 否则按组合找到的记录整行更新
    返回统计 {created, updated, skipped, errors}
    （skipped 恒为 0，保留字段仅兼容旧调用）
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {'created': 0, 'updated': 0, 'skipped': 0, 'errors': ['文件为空']}
    header_cells = all_rows[0]
    col_map = _match_columns(header_cells)
    if not col_map:
        return {'created': 0, 'updated': 0, 'skipped': 0,
                'errors': ['未识别到任何已知列（名称/平台/接口地址/API Key 等）']}

    data_rows = all_rows[1:]
    stats = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

    conn = db.get_conn()
    try:
        max_order = conn.execute("SELECT MAX(sort_order) AS m FROM key_vault").fetchone()["m"] or 0
        for r_i, row in enumerate(data_rows, start=2):
            rec = {}
            for c_idx, field in col_map.items():
                val = row[c_idx] if c_idx < len(row) else None
                rec[field] = '' if val is None else str(val).strip()
            name = rec.get('name')
            if not name:
                stats['errors'].append(f'第{r_i}行：缺少「名称」，已跳过')
                continue

            provider = rec.get('provider') or ''
            account = rec.get('account') or ''
            existing = conn.execute(
                "SELECT id FROM key_vault WHERE name = ? AND provider = ? AND account = ?",
                (name, provider, account)
            ).fetchone()

            if existing:
                # 按组合键更新整行
                vid = existing['id']
                sets, vals = [], []
                for f in FIELDS:
                    sets.append(f"{f} = ?")
                    vals.append(rec.get(f) or '')
                sets.append("updated_at = datetime('now')")
                conn.execute(
                    f"UPDATE key_vault SET {','.join(sets)} WHERE id = ?", vals + [vid])
                stats['updated'] += 1
                continue

            # 新建
            cols, vals = [], []
            for f in FIELDS:
                cols.append(f)
                v = rec.get(f)
                vals.append(v if v is not None else '')
            max_order += 10
            cols.append('sort_order')
            vals.append(max_order)
            ph = ','.join('?' * len(cols))
            conn.execute(
                f"INSERT INTO key_vault ({','.join(cols)}) VALUES ({ph})", vals)
            stats['created'] += 1
        conn.commit()
    except Exception as e:
        conn.rollback()
        stats['errors'].append(f'导入失败：{e}')
    finally:
        conn.close()
    return stats
